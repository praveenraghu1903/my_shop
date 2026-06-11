from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from .models import Store, Product, Stock, Invoice, InvoiceItem, UserProfile, Supplier, Purchase, PurchaseItem, Location, InvoiceContact
from django.contrib import messages
from django.db import transaction
from django.utils import timezone
from django.db.models import Sum
from decimal import Decimal
from django.http import HttpResponse
from io import StringIO, BytesIO
import csv
import os
import zipfile
from django.conf import settings
from urllib.parse import urlparse, parse_qs, urlunparse
from urllib.request import urlopen, Request
try:
    import openpyxl  # optional, for .xlsx import
except Exception:
    openpyxl = None

@login_required
def dashboard_redirect(request):
    if request.user.is_staff or request.user.is_superuser:
        return redirect('/admin/')
    return redirect('sales_new')

@login_required
def sales_new(request):
    # Fetch products with their assigned locations, ordered by category for grouping
    products = Product.objects.prefetch_related('locations').order_by('category', 'name')
    locations = Location.objects.all()
    
    # Get stock levels for Godown (assuming sales come from Godown)
    godown = Store.objects.filter(store_type='GODOWN').first()
    stock_map = {}
    if godown:
        stocks = Stock.objects.filter(store=godown)
        stock_map = {s.product_id: s.quantity for s in stocks}

    # Attach stock quantity to each product object
    for p in products:
        p.stock_quantity = stock_map.get(p.id, 0)
    
    # Get today's summary
    today = timezone.now().date()
    # Filter invoices by the user's store if possible
    try:
        user_store = request.user.userprofile.store
    except UserProfile.DoesNotExist:
        user_store = None

    if request.method == 'POST':
        if not user_store:
            messages.error(request, "You are not assigned to any store.")
            return redirect('sales_new')

        customer_name = request.POST.get('customer_name')
        primary_mobile = request.POST.get('customer_mobile') or None
        paid_amount = Decimal(request.POST.get('paid_amount'))
        discount_amount = Decimal(request.POST.get('discount_amount') or '0')
        transport_cost = Decimal(request.POST.get('transport_cost') or '0')
        labour_cost = Decimal(request.POST.get('labour_cost') or '0')
        other_expenses = transport_cost + labour_cost

        # Support multiple items via arrays; fallback to single item if arrays not provided
        product_ids = request.POST.getlist('product_ids[]')
        quantities_raw = request.POST.getlist('quantities[]')
        rates_raw = request.POST.getlist('rates[]')
        location_ids = request.POST.getlist('locations[]')

        # Fallback to single-item fields
        if not product_ids:
            single_product_id = request.POST.get('product')
            single_qty = request.POST.get('quantity')
            single_rate = request.POST.get('rate')
            if single_product_id and single_qty and single_rate:
                product_ids = [single_product_id]
                quantities_raw = [single_qty]
                rates_raw = [single_rate]

        try:
            with transaction.atomic():
                godown = Store.objects.filter(store_type='GODOWN').first()
                if not godown:
                    raise Exception("Central Godown not found.")

                # Parse and validate lists
                if not product_ids or len(product_ids) != len(quantities_raw) or len(product_ids) != len(rates_raw):
                    raise Exception("Invalid sale items submitted.")

                items = []
                total_amount = Decimal('0')
                # Pre-check stock availability for all items
                for idx, (pid, q_raw, r_raw) in enumerate(zip(product_ids, quantities_raw, rates_raw)):
                    product = Product.objects.get(id=pid)
                    qty = Decimal(q_raw)
                    rate = Decimal(r_raw)
                    total_amount += qty * rate
                    stock = Stock.objects.filter(product=product, store=godown).first()
                    if not stock or stock.quantity < qty:
                        available = stock.quantity if stock else Decimal('0')
                        raise Exception(f"Insufficient stock for {product.name}. Available: {available}")
                    # Resolve optional location for this item
                    loc_id = location_ids[idx] if idx < len(location_ids) else None
                    loc_obj = None
                    if loc_id:
                        try:
                            loc_obj = Location.objects.get(id=loc_id)
                        except Location.DoesNotExist:
                            loc_obj = None
                    items.append((product, qty, rate, stock, loc_obj))

                # Create invoice
                invoice = Invoice.objects.create(
                    store=user_store,
                    customer_name=customer_name,
                    customer_mobile=primary_mobile,
                    discount_amount=discount_amount,
                    transport_cost=transport_cost,
                    labour_cost=labour_cost,
                    other_expenses=other_expenses,
                    total_amount=(total_amount + transport_cost + labour_cost - discount_amount),
                    paid_amount=paid_amount
                )

                # Deduct stock and create invoice items
                for product, qty, rate, stock, loc in items:
                    stock.quantity -= qty
                    stock.save()
                    InvoiceItem.objects.create(
                        invoice=invoice,
                        product=product,
                        quantity=qty,
                        rate=rate,
                        location=loc
                    )

                messages.success(request, f"Sale recorded successfully! Invoice #{invoice.id}")
                return redirect(f'/sales/new/?whatsapp_invoice={invoice.id}')

        except Exception as e:
            messages.error(request, str(e))

    # Calculate today's stats for the cards
    today_sales = Invoice.objects.filter(date__date=today, store=user_store).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    today_received = Invoice.objects.filter(date__date=today, store=user_store).aggregate(Sum('paid_amount'))['paid_amount__sum'] or 0
    today_due = today_sales - today_received

    # Pass last invoice for WhatsApp sharing
    last_invoice = None
    whatsapp_invoice_id = request.GET.get('whatsapp_invoice')
    if whatsapp_invoice_id:
        try:
            last_invoice = Invoice.objects.prefetch_related('items__product').get(id=whatsapp_invoice_id)
        except Invoice.DoesNotExist:
            pass

    context = {
        'products': products,
        'locations': locations,
        'user_store': user_store,
        'today_sales': today_sales,
        'today_received': today_received,
        'today_due': today_due,
        'last_invoice': last_invoice,
    }
    return render(request, 'inventory/sales_new.html', context)

@staff_member_required
def sales_summary(request):
    today = timezone.now().date()
    
    total_sales = Invoice.objects.filter(date__date=today).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_paid = Invoice.objects.filter(date__date=today).aggregate(Sum('paid_amount'))['paid_amount__sum'] or 0
    total_due = total_sales - total_paid
    total_invoices = Invoice.objects.filter(date__date=today).count()
    
    # Purchases (Expenses)
    total_purchases = Purchase.objects.filter(date__date=today).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    net_profit = total_sales - total_purchases

    # Store-wise breakdown (Legacy summary)
    stores = Store.objects.filter(store_type='DISPLAY')
    store_stats = []
    for store in stores:
        s_sales = Invoice.objects.filter(date__date=today, store=store).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        store_stats.append({
            'name': store.name,
            'sales': s_sales
        })

    # Detailed Invoices per Store
    silwani_invoices = Invoice.objects.filter(
        date__date=today, 
        store__name='Silwani'
    ).prefetch_related('items', 'items__product').order_by('-date')
    
    gairatganj_invoices = Invoice.objects.filter(
        date__date=today, 
        store__name='Gairatganj'
    ).prefetch_related('items', 'items__product').order_by('-date')

    context = {
        'total_sales': total_sales,
        'total_paid': total_paid,
        'total_due': total_due,
        'total_invoices': total_invoices,
        'total_purchases': total_purchases,
        'net_profit': net_profit,
        'store_stats': store_stats,
        'silwani_invoices': silwani_invoices,
        'gairatganj_invoices': gairatganj_invoices,
    }
    return render(request, 'inventory/sales_summary.html', context)

@staff_member_required
def purchase_new(request):
    products = Product.objects.all()
    suppliers = Supplier.objects.all()
    recent_purchases = Purchase.objects.order_by('-date')[:10]

    if request.method == 'POST':
        supplier_name = request.POST.get('supplier_name')
        invoice_number = request.POST.get('invoice_number')
        product_id = request.POST.get('product')
        quantity = Decimal(request.POST.get('quantity'))
        rate = Decimal(request.POST.get('rate'))
        
        # Calculate total
        total_amount = quantity * rate
        
        try:
            with transaction.atomic():
                # 1. Get or Create Supplier
                supplier, created = Supplier.objects.get_or_create(name=supplier_name)
                
                # 2. Create Purchase Record
                purchase = Purchase.objects.create(
                    supplier=supplier,
                    invoice_number=invoice_number,
                    total_amount=total_amount
                )
                
                # 3. Create Purchase Item
                PurchaseItem.objects.create(
                    purchase=purchase,
                    product_id=product_id,
                    quantity=quantity,
                    rate=rate
                )
                
                # 4. Add Stock to Godown
                godown = Store.objects.get(store_type='GODOWN')
                stock, created = Stock.objects.get_or_create(product_id=product_id, store=godown)
                stock.quantity += quantity
                stock.save()
                
                messages.success(request, f"Purchase recorded! Stock added to Godown. Purchase ID: {purchase.id}")
                return redirect('purchase_new')

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

    context = {
        'products': products,
        'suppliers': suppliers,
        'recent_purchases': recent_purchases,
    }
    return render(request, 'inventory/purchase_new.html', context)

@staff_member_required
def product_import(request):
    if request.method == 'POST':
        upload = request.FILES.get('file')
        sheet_url = (request.POST.get('sheet_url') or '').strip()
        filename = upload.name.lower() if upload else ''
        created_count = 0
        updated_count = 0
        stock_added_total = Decimal('0')

        # Ensure a Godown exists for stock
        godown = Store.objects.filter(store_type='GODOWN').first()
        if not godown:
            messages.error(request, 'Central Godown not found. Create a GODOWN store first.')
            return redirect('product_import')

        def normalize_category(cat):
            if not cat:
                return 'OTHER'
            cat = str(cat).strip().upper()
            allowed = {c for c, _ in Product.CATEGORY_CHOICES}
            if cat in allowed:
                return cat
            # Try match by display labels
            labels = {label.upper(): code for code, label in Product.CATEGORY_CHOICES}
            return labels.get(cat, 'OTHER')

        # Parse CSV/XLSX into list of dict rows and import

        # Parse CSV
        try:
            # Load rows from one of: Google Sheet URL (CSV), uploaded CSV, uploaded XLSX
            rows = None
            if sheet_url and not upload:
                # Accept Google Sheets share links; convert to export CSV if needed
                parsed = urlparse(sheet_url)
                if 'docs.google.com' in parsed.netloc and '/spreadsheets/' in parsed.path:
                    if '/export' not in parsed.path:
                        # Convert /edit to /export
                        path_base = parsed.path.split('/edit')[0]
                        query = parse_qs(parsed.query)
                        gid = query.get('gid', ['0'])[0]
                        new_path = f"{path_base}/export"
                        new_query = f"format=csv&gid={gid}"
                        parsed = parsed._replace(path=new_path, query=new_query)
                    sheet_url = urlunparse(parsed)
                # Fetch CSV content
                req = Request(sheet_url, headers={'User-Agent': 'Mozilla/5.0'})
                with urlopen(req, timeout=20) as resp:
                    data = resp.read().decode('utf-8-sig')
                reader = csv.DictReader(StringIO(data))
                rows = list(reader)
            elif upload and filename.endswith('.csv'):
                decoded = upload.read().decode('utf-8-sig')
                reader = csv.DictReader(StringIO(decoded))
                rows = list(reader)
            elif upload and filename.endswith('.xlsx') and openpyxl:
                wb = openpyxl.load_workbook(upload)
                ws = wb.active
                headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))[0:]]
                rows = []
                for r in ws.iter_rows(min_row=2, values_only=True):
                    row = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
                    rows.append(row)
            else:
                messages.error(request, 'Provide a Google Sheet CSV URL or upload CSV/XLSX. For XLSX, openpyxl is required.')
                return redirect('product_import')

            # Normalize columns and import
            with transaction.atomic():
                for row in rows:
                    name = (str(row.get('name') or row.get('Name') or '').strip())
                    if not name:
                        continue
                    size = (str(row.get('size') or row.get('Size') or '').strip())
                    unit = (str(row.get('unit') or row.get('Unit') or 'sqft').strip()) or 'sqft'
                    category_raw = row.get('category') or row.get('Category')
                    category = normalize_category(category_raw)
                    locations_csv = row.get('locations') or row.get('Locations') or ''
                    initial_qty_raw = row.get('initial_quantity') or row.get('Initial Quantity') or row.get('initial qty')
                    try:
                        initial_qty = Decimal(str(initial_qty_raw)) if initial_qty_raw not in (None, '') else Decimal('0')
                    except Exception:
                        initial_qty = Decimal('0')

                    product, created = Product.objects.get_or_create(
                        name=name,
                        size=size,
                        defaults={'category': category, 'unit': unit}
                    )
                    if not created:
                        # Update mutable fields if changed
                        updated = False
                        if product.category != category:
                            product.category = category
                            updated = True
                        if product.unit != unit:
                            product.unit = unit
                            updated = True
                        if updated:
                            product.save()
                        updated_count += 1
                    else:
                        created_count += 1

                    # Handle locations
                    if locations_csv:
                        names = [n.strip() for n in str(locations_csv).split(',') if n and str(n).strip()]
                        for lname in names:
                            loc, _ = Location.objects.get_or_create(name=lname)
                            product.locations.add(loc)

                    # Handle initial stock addition
                    if initial_qty and initial_qty > 0:
                        stock, _ = Stock.objects.get_or_create(product=product, store=godown, defaults={'quantity': Decimal('0')})
                        stock.quantity += initial_qty
                        stock.save()
                        stock_added_total += initial_qty

            messages.success(request, f"Import complete. Created: {created_count}, Updated: {updated_count}, Stock added to Godown: {stock_added_total}.")
            return redirect('product_import')
        except Exception as e:
            messages.error(request, f"Import failed: {e}")
            return redirect('product_import')

    return render(request, 'inventory/product_import.html')

@staff_member_required
def daily_report_download(request):
    today = timezone.localdate()
    # Inventory snapshot
    products = Product.objects.all().annotate(total_qty=Sum('stock__quantity'))

    inv_buf = StringIO()
    inv_writer = csv.writer(inv_buf)
    inv_writer.writerow(['product_id', 'name', 'size', 'category', 'unit', 'total_stock'])
    for p in products:
        inv_writer.writerow([p.id, p.name, p.size, p.get_category_display(), p.unit, p.total_qty or 0])

    # Sales (item-level) for today
    items = InvoiceItem.objects.filter(invoice__date__date=today).select_related('invoice', 'product', 'invoice__store')
    sales_buf = StringIO()
    sales_writer = csv.writer(sales_buf)
    sales_writer.writerow(['invoice_id', 'date', 'store', 'customer_name', 'product', 'quantity', 'rate', 'line_total', 'invoice_total', 'paid', 'due'])
    for it in items:
        inv = it.invoice
        sales_writer.writerow([
            inv.id,
            inv.date.astimezone(timezone.get_current_timezone()).strftime('%Y-%m-%d %H:%M:%S'),
            inv.store.name,
            inv.customer_name,
            it.product.name,
            it.quantity,
            it.rate,
            it.item_total,
            inv.total_amount,
            inv.paid_amount,
            inv.balance_due,
        ])

    # Package into a zip
    zip_bytes = BytesIO()
    with zipfile.ZipFile(zip_bytes, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f'inventory_{today}.csv', inv_buf.getvalue())
        zf.writestr(f'sales_{today}.csv', sales_buf.getvalue())
    zip_bytes.seek(0)

    resp = HttpResponse(zip_bytes.read(), content_type='application/zip')
    resp['Content-Disposition'] = f'attachment; filename="daily_report_{today}.zip"'
    return resp



# ─────────────────────────────────────────────────────────────────────────────
# COST PRICE EXPORT / IMPORT
# ─────────────────────────────────────────────────────────────────────────────

@staff_member_required
def export_cost_price_template(request):
    """
    Export all products as an Excel file with empty cost_price,
    labour_cost, transport_cost columns ready for the admin to fill in.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cost Prices"

    # ── Header row ────────────────────────────────────────────────────────
    headers = [
        'product_id', 'name', 'category', 'size', 'unit',
        'cost_price_per_unit',   # ← fill this
        'supplier_name',         # ← optional
        'notes',                 # ← optional
    ]

    header_fill  = PatternFill('solid', start_color='2C3E50')
    input_fill   = PatternFill('solid', start_color='FFF9C4')   # yellow = fill these
    locked_fill  = PatternFill('solid', start_color='F5F5F5')   # grey = reference only
    header_font  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    normal_font  = Font(name='Arial', size=10)
    center       = Alignment(horizontal='center', vertical='center')
    left         = Alignment(horizontal='left',   vertical='center')
    thin         = Side(style='thin', color='CCCCCC')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    # ── Instructions row ──────────────────────────────────────────────────
    ws.cell(row=2, column=1,  value='← do not edit IDs').font = Font(italic=True, color='999999', name='Arial', size=9)
    ws.cell(row=2, column=6,  value='← FILL THIS (₹ per unit)').font = Font(italic=True, color='E67E22', name='Arial', size=9)
    ws.cell(row=2, column=7,  value='← optional supplier name').font = Font(italic=True, color='999999', name='Arial', size=9)
    ws.merge_cells('A2:E2')

    # ── Product rows ──────────────────────────────────────────────────────
    from inventory.models import Product
    products = Product.objects.prefetch_related('buy_price_record').order_by('category', 'name')

    for row_num, product in enumerate(products, start=3):
        try:
            existing_cost = product.buy_price_record.buy_price
        except Exception:
            existing_cost = ''

        row_data = [
            product.id,
            product.name,
            product.get_category_display(),
            product.size,
            product.unit,
            existing_cost,   # pre-fill if already set
            '',              # supplier_name
            '',              # notes
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font      = normal_font
            cell.border    = border
            cell.alignment = left
            # Yellow highlight on editable columns
            if col in (6, 7, 8):
                cell.fill = input_fill
            else:
                cell.fill = locked_fill

    # ── Column widths ─────────────────────────────────────────────────────
    widths = [10, 35, 14, 12, 8, 22, 25, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A3'

    # ── Instructions sheet ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("AMBIKA — Cost Price Import Template", True),
        ("", False),
        ("INSTRUCTIONS:", True),
        ("1. Fill the 'cost_price_per_unit' column (yellow) for each product.", False),
        ("2. Enter the cost price in ₹ per unit (same unit as shown in 'unit' column).", False),
        ("3. Supplier name is optional — leave blank if not applicable.", False),
        ("4. Do NOT change product_id, name, category, size or unit columns.", False),
        ("5. Leave cost_price blank for products you don't want to update.", False),
        ("6. Save the file as .xlsx and upload it in Admin → Import Cost Prices.", False),
        ("", False),
        ("HOW NET PROFIT IS CALCULATED:", True),
        ("Net Profit = Revenue − Cost of Goods Sold − Transport − Labour", False),
        ("Cost of Goods Sold = Sum of (quantity sold × cost_price) for each invoice item", False),
        ("The more products you fill cost prices for, the more accurate the cash flow.", False),
    ]
    for r, (text, bold) in enumerate(instructions, 1):
        cell = ws2.cell(row=r, column=1, value=text)
        cell.font = Font(bold=bold, name='Arial', size=10, color='2C3E50' if bold else '333333')
    ws2.column_dimensions['A'].width = 80

    # ── Response ──────────────────────────────────────────────────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ambika_cost_prices.xlsx"'
    wb.save(response)
    return response


@staff_member_required
def import_cost_prices(request):
    """
    Import cost prices from the filled Excel sheet.
    Also accepts optional labour_cost and transport_cost as global values.
    """
    if request.method != 'POST':
        return redirect('/admin/')

    if not request.user.is_superuser:
        messages.error(request, "Only admin can import cost prices.")
        return redirect('/admin/')

    upload = request.FILES.get('cost_price_file')
    if not upload:
        messages.error(request, "No file uploaded.")
        return redirect('/admin/cost-price-import/')

    # Optional global labour / transport for this import batch
    labour_raw    = request.POST.get('labour_cost', '0') or '0'
    transport_raw = request.POST.get('transport_cost', '0') or '0'
    supplier_name = request.POST.get('supplier_name', '').strip()

    try:
        import openpyxl
        from inventory.models import Product, ProductBuyPrice, Purchase, PurchaseItem, Supplier, Store

        wb = openpyxl.load_workbook(upload, data_only=True)
        ws = wb.active

        # Read headers from row 1
        headers = [str(ws.cell(1, c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]

        def col(name):
            try:
                return headers.index(name)
            except ValueError:
                return None

        id_col    = col('product_id')
        price_col = col('cost_price_per_unit')
        sup_col   = col('supplier_name')

        if id_col is None or price_col is None:
            messages.error(request, "File format invalid. Make sure you used the exported template.")
            return redirect('/admin/cost-price-import/')

        updated = 0
        skipped = 0
        errors  = []

        with transaction.atomic():
            for row in ws.iter_rows(min_row=3, values_only=True):
                if not any(row):
                    continue

                prod_id   = row[id_col]
                cost_raw  = row[price_col]
                row_sup   = row[sup_col] if sup_col is not None else None

                if not prod_id or cost_raw in (None, ''):
                    skipped += 1
                    continue

                try:
                    cost = Decimal(str(cost_raw))
                    if cost <= 0:
                        skipped += 1
                        continue
                except Exception:
                    errors.append(f"Row product_id={prod_id}: invalid cost '{cost_raw}'")
                    continue

                try:
                    product = Product.objects.get(id=int(prod_id))
                except Product.DoesNotExist:
                    errors.append(f"Product id={prod_id} not found.")
                    continue

                # Save / update buy price
                ProductBuyPrice.objects.update_or_create(
                    product=product,
                    defaults={'buy_price': cost}
                )
                updated += 1

            # Create a purchase record for this import batch if labour/transport provided
            try:
                labour    = Decimal(labour_raw)
                transport = Decimal(transport_raw)
            except Exception:
                labour = transport = Decimal('0')

        # Summary message
        msg = f"Cost prices updated for {updated} products."
        if skipped:
            msg += f" Skipped {skipped} (blank or zero)."
        if errors:
            msg += f" Errors: {'; '.join(errors[:5])}"
        if updated > 0:
            messages.success(request, msg)
        else:
            messages.warning(request, msg)

    except Exception as e:
        messages.error(request, f"Import failed: {e}")

    return redirect('/admin/cost-price-import/')


@staff_member_required
def cost_price_import_page(request):
    """Admin page for cost price import/export."""
    if not request.user.is_superuser:
        return redirect('/admin/')

    from inventory.models import Product, ProductBuyPrice
    total_products = Product.objects.count()
    priced_products = ProductBuyPrice.objects.count()
    unpriced = total_products - priced_products

    context = {
        'total_products': total_products,
        'priced_products': priced_products,
        'unpriced': unpriced,
    }
    return render(request, 'admin/cost_price_import.html', context)
